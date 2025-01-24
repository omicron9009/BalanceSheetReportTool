import openpyxl
from openpyxl.styles import Border, Side, Font, Alignment

# Create a new workbook and select the active worksheet
output_wb = openpyxl.Workbook()
output_ws = output_wb.active

# Take input for name and address
company_name = "Saptaranga Research and Organic Private Limited"
address_line1 = "Plot No 45,"
address_line2 = "Ravindra Nagar P.M.G. Society"
address_line3 = "NOTES ANNEXED TO AND FORMING PART OF ACCOUNTS FOR THE YEAR ENDING 31-Mar-2023"
dateCurr = "As on 31-Mar-2023"
datePrev = "As on 31-Mar-2022"

# Set the title of the sheet
output_ws.title = "Financial Report"

# Set the title and subtitle
output_ws.merge_cells('C1:H1')
output_ws['C1'] = company_name
output_ws.merge_cells('C2:H2')
output_ws['C2'] = address_line1
output_ws.merge_cells('C3:H3')
output_ws['C3'] = address_line2
output_ws.merge_cells('C4:H4')
output_ws['C4'] = address_line3

# Apply formatting
title_font = Font(size=14, bold=True)
header_font = Font(size=10, bold=True)

for cell in ['C1', 'C2', 'C3', 'C4']:
    output_ws[cell].font = title_font
    output_ws[cell].alignment = Alignment(horizontal='center')

# Adjust column widths
output_ws.column_dimensions['C'].width = 50
output_ws.column_dimensions['G'].width = 15
output_ws.column_dimensions['H'].width = 15

# Define the border style
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Apply thin border to each cell in the title and subtitle section
for row in range(3, 5):
    for col in range(3, 9):  # C to H
        cell = output_ws.cell(row=row, column=col)
        cell.border = thin_border

# Define the thick border style
thick_border = Border(
    left=Side(style='thick'),
    right=Side(style='thick'),
    top=Side(style='thick'),
    bottom=Side(style='thick')
)

def add_section(headers, search_head, note, output_ws, row_offset):
    # Load the input Excel file
    input_wb = openpyxl.load_workbook("C:\\Users\\jadit\\OneDrive\\Desktop\\tribalop\\ManyParticularsOutput.xlsx")
    input_ws = input_wb.active
    ws = input_ws

    # Set the headers
    output_ws[f'C{6 + row_offset}'] = note
    output_ws[f'G{5 + row_offset}'] = "In Rs. hundreds"
    output_ws[f'C{7 + row_offset + len(headers)}'] = "Total"

    output_ws[f'G{6 + row_offset}'] = dateCurr
    output_ws[f'H{6 + row_offset}'] = datePrev

    output_ws[f'G{5 + row_offset}'].font = header_font
    output_ws[f'G{6 + row_offset}'].font = header_font
    output_ws[f'H{6 + row_offset}'].font = header_font
    output_ws[f'C{6 + row_offset}'].font = header_font
    output_ws[f'C{7 + row_offset + len(headers)}'].font=header_font

    # Apply the border to the header cells
    for cell in [f'C{6 + row_offset}', f'G{5 + row_offset}', f'G{6 + row_offset}', f'H{6 + row_offset}']:
        output_ws[cell].border = thin_border

    # Place each header in a separate row
    for i, header in enumerate(headers):
        output_ws[f'C{7 + row_offset + i}'] = header
        output_ws[f'C{7 + row_offset + i}'].border = thin_border

    # Specify the header name of the column you want to iterate through
    header_name = "Major Head"
    subHeadList = []
    mappedDict = {}

    # Find the index of the column with the specified header name
    column_index = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == header_name:
            column_index = col
            break

    if column_index is not None:
        # Iterate through each row in the column
        for row in range(2, ws.max_row + 1):  # Start from 2 to skip the header row
            cell_value = ws.cell(row=row, column=column_index).value
            if cell_value == search_head:
                subHeadList.append(ws.cell(row=row, column=column_index + 1).value)

        for row in range(2, ws.max_row + 1):
            sub_head_value = ws.cell(row=row, column=column_index + 1).value
            if sub_head_value in subHeadList:
                mappedDict[sub_head_value] = [
                    ws.cell(row=row, column=column_index - 2).value,
                    ws.cell(row=row, column=column_index - 1).value
                ]

    sumOfcurr = sum(mappedDict[key][0] for key in mappedDict if mappedDict[key][0] is not None)
    sumofprev = sum(mappedDict[key][1] for key in mappedDict if mappedDict[key][1] is not None)

    for c in range(1, output_ws.max_column + 1):
        for r in range(1, output_ws.max_row + 1):
            cell_value = output_ws.cell(row=r, column=c).value
            if cell_value in mappedDict:
                output_ws.cell(row=r, column=c + 4).value = mappedDict[cell_value][0]
                output_ws.cell(row=r, column=c + 5).value = mappedDict[cell_value][1]
            if cell_value == "Total" and (r >= (7 + row_offset) and r <= (7 + row_offset + len(headers))):
                output_ws.cell(row=r, column=c + 4).value = sumOfcurr
                output_ws.cell(row=r, column=c + 5).value = sumofprev

    # Apply thick border to the row immediately after the section
    for col in range(3, 9):  # C to H
        cell = output_ws.cell(row=7 + row_offset + len(headers) + 1, column=col)
        cell.border = thick_border

    input_wb.close()

# Define the sections to be added
otherCurrentLiabilities = [
    "Current Maturities Of Finance Lease Obligations",
    "Interest Accrued But Not Due On Borrowings",
    "Interest Accrued And Due On Borrowings",
    "Income Received In Advance",
    "Unpaid Dividends",
    "Application Money Received For Allotment Of Securities And Due For Refund And Interest Accrued Thereon",
    "Unpaid Matured Deposits And Interest Accrued Thereon",
    "Unpaid Matured Debentures And Interest Accrued Thereon",
    "Other Payables (Specify Nature)"
]
DeferredTaxLiability = ["Deferred Tax Liability"]
longtermliabilities=["Trade Payables", "Others"]
cashandcashequivalent=[
    "Cheques, Drafts On Hand",
    "Cash On Hand",
    "Cash At Bank",
    "Other C & CE (Specify Nature)"
]
CurrentInvestments=[
    "Investments In Equity Instruments",
    "Investments In Preference Shares",
    "Investments In Government Or Trust Securities",
    "Investments In Debentures Or Bonds",
    "Investments In Mutual Funds",
    "Investments In Partnership Firms",
    "Other Investments (Specify Nature)"
]
LongtermProvisions=[
    "Provision For Employee Benefits",
    "Others Long Term Provision (Specify Nature)"
]
shortTermProvisions=[
    "Provision for Employees Benefits",
    "Provision for Income Tax",
    "Other Short Term Provision (Specify Nature)"
]
nonCurrentInvestments=[
    "Investment Property",
    "Investments In Equity Instruments",
    "Investments In Preference Shares",
    "Investments In Government Or Trust Securities",
    "Investments In Debentures Or Bonds",
    "Investments In Mutual Funds",
    "Investments In Partnership Firms",
    "Other Non-Current Investments (Specify Nature)"
]
deferredtaxassets=["Deferred Tax Asset"]
inventories=[
    "Raw Materials",
    "Work-In-Progress",
    "Finished Goods",
    "Stock-In-Trade",
    "Stores And Spares",
    "Loose Tools",
    "Other Inventory (Specify Nature)"
]
sections = [
    {"header": DeferredTaxLiability, "search_head": "Deferred Tax Liability", "note": "Note : Deferred Tax Liability"},
    {"header": longtermliabilities, "search_head": "Other Long-term Liabilities", "note": "Note : Other Long-term Liabilities"},
    {"header": LongtermProvisions, "search_head": "Long-Term Provisions", "note": "Note : Long-Term Provisions"},
    {"header": otherCurrentLiabilities, "search_head": "Other Current Liabilities", "note": "Note : Other Current Liabilities"},    
    {"header": shortTermProvisions, "search_head": "Short Term Provisions", "note": "Note : Short Term Provisions"},
    {"header": nonCurrentInvestments, "search_head": "Non Current Investments", "note": "Note : Non Current Investments"},
    {"header": deferredtaxassets, "search_head": "Deferred Tax Asset", "note": "Note : Deferred Tax Asset"},
    {"header": CurrentInvestments, "search_head": "Current Investments", "note": "Note : Current Investments"},

    #{"header": inventories, "search_head": "Inventories", "note": "Note : Inventories"},


    {"header": cashandcashequivalent, "search_head": "Cash and Cash Equivalents", "note": "Note : Cash and Cash Equivalents"}


]
# Add each section to the worksheet
row_offset = 0
for section in sections:
    headers = section["header"] if isinstance(section["header"], list) else [section["header"]]
    add_section(headers, section["search_head"], section["note"], output_ws, row_offset)
    row_offset += 3 + len(headers)  # Adjust this value based on the number of rows each section takes

# Apply thin border to all content cells
for row in output_ws.iter_rows(min_row=1, max_row=output_ws.max_row, min_col=3, max_col=output_ws.max_column):
    for cell in row:
        cell.border = thin_border

# Save the final workbook
output_wb.save("C:\\Users\\jadit\\OneDrive\\Desktop\\BalanceSheetProject\\GeneratedOutput\\TemplateOutput\\CombinedFinancialReportManyInputProblemSol.xlsx")

