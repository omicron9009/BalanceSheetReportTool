#Libraries
import openpyxl
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl import load_workbook
# Create a new workbook and select the active worksheet
output_wb = openpyxl.Workbook()
bs_ws = output_wb.active
bs_ws.title = "BS Notes"
# Define a thin border style

# Create another worksheet for PL notes
pl_ws = output_wb.create_sheet(title="PL Notes")
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
# Add Details For the Company
company_name = "Saptaranga Research and Organic Private Limited"
address_line1 = "Plot No 45,"
address_line2 = "Ravindra Nagar P.M.G. Society"
address_line3 = "NOTES ANNEXED TO AND FORMING PART OF ACCOUNTS FOR THE YEAR ENDING 31-Mar-2023"
dateCurr = "As on 31-Mar-2023"
datePrev = "As on 31-Mar-2022"
inputPath="C:\\Users\\jadit\\OneDrive\\Desktop\\BalanceSheetProject\\ExcelFiles\\InputForScheduleTemplate\\InputForASchedule.xlsx"
outputPath="C:\\Users\\jadit\\OneDrive\\Desktop\\BalanceSheetProject\\GeneratedOutput\\TemplateOutput\\Schedules-distributed-BS-PL-Format-3.xlsx"
# inputPath=input("Enter the Path of the Grouped Trial Balance : ")
# outputPath=input("Enter the path where you want to save the Schedules with file *NAME* :")
# company_name=input("Enter Company Name :")
# address_line1=input("Enter address line 1 :")
# address_line2=input("Address line 2 : ")
# address_line3=input("Address Line 3 : ")
# dateCurr=input("Enter Current Year date, in format 31-Mar-2023 : ")
# datePrev=input("Enter Previous Year date, in format 31-Mar-2023 : ")
# Function to set the title and address in a worksheet
def set_title_and_address(ws):
    ws.merge_cells('C1:H1')
    ws['C1'] = company_name
    ws.merge_cells('C2:H2')
    ws['C2'] = address_line1
    ws.merge_cells('C3:H3')
    ws['C3'] = address_line2
    ws.merge_cells('C4:H4')
    ws['C4'] = address_line3

    title_font = Font(size=14, bold=True)
    for cell in ['C1', 'C2', 'C3', 'C4']:
        ws[cell].font = title_font
        ws[cell].alignment = Alignment(horizontal='center')

    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15

# Apply title and address to both sheets
set_title_and_address(bs_ws)
set_title_and_address(pl_ws)

# Define the border styles
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

#-----------------------------------------------addd section----Only Sub-Head-----------------------------------
def add_section(headers, search_head, note, category, row_offset):
    # Select the appropriate worksheet based on category
    output_ws = bs_ws if category == 'BS' else pl_ws

    # Load the input Excel file
    input_wb = openpyxl.load_workbook(inputPath)
    input_ws = input_wb.active

    # Set the headers
    output_ws[f'C{6 + row_offset}'] = note
    output_ws[f'C{6 + row_offset}'].font=Font(bold=True)
    output_ws[f'H{5 + row_offset}'] = "In Rs. hundreds"
    output_ws[f'C{7 + row_offset + len(headers)}'] = "Total"
    output_ws[f'C{7 + row_offset + len(headers)}'].font=Font(bold=True)
    output_ws[f'G{7 + row_offset + len(headers)}'].font=Font(bold=True)
    output_ws[f'H{7 + row_offset + len(headers)}'].font=Font(bold=True)

    output_ws[f'G{6 + row_offset}'] = dateCurr
    output_ws[f'H{6 + row_offset}'] = datePrev

    header_font = Font(size=10, bold=True)
    output_ws[f'G{5 + row_offset}'].font = header_font
    output_ws[f'G{6 + row_offset}'].font = header_font
    output_ws[f'H{6 + row_offset}'].font = header_font
    output_ws[f'C{6 + row_offset}'].font = header_font
    output_ws[f'C{7 + row_offset + len(headers)}'].font = header_font

    # Apply the border to the header cells
    # for cell in [f'C{6 + row_offset}', f'G{5 + row_offset}', f'G{6 + row_offset}', f'H{6 + row_offset}']:
    #     output_ws[cell].border = thin_border

    # Place each header in a separate row
    for i, header in enumerate(headers):
        output_ws[f'C{7 + row_offset + i}'] = header
        # output_ws[f'C{7 + row_offset + i}'].border = thin_border

    # Specify the header name of the column you want to iterate through
    header_name = "Major Head"
    subHeadList = []
    mappedDict = {}

    # Find the index of the column with the specified header name
    column_index = None
    for col in range(1, input_ws.max_column + 1):
        if input_ws.cell(row=1, column=col).value == header_name:
            column_index = col
            break

    if column_index is not None:
        # Iterate through each row in the column
        for row in range(2, input_ws.max_row + 1):  # Start from 2 to skip the header row
            cell_value = input_ws.cell(row=row, column=column_index).value
            if cell_value == search_head:
                subHeadList.append(input_ws.cell(row=row, column=column_index + 1).value)

        for row in range(2, input_ws.max_row + 1):
            sub_head_value = input_ws.cell(row=row, column=column_index + 1).value
            currYearval=input_ws.cell(row=row, column=column_index - 2).value or 0 
            prevYearval=input_ws.cell(row=row, column=column_index - 1).value or 0
            if sub_head_value in subHeadList:
                if(sub_head_value in mappedDict):
                    mappedDict[sub_head_value][0]+=currYearval
                    mappedDict[sub_head_value][1]+=prevYearval
                else :
                    mappedDict[sub_head_value] = [currYearval,prevYearval]

    sumOfcurr = sum(mappedDict[key][0] for key in mappedDict if mappedDict[key][0] is not None)
    sumofprev = sum(mappedDict[key][1] for key in mappedDict if mappedDict[key][1] is not None)

    for c in range(1, output_ws.max_column + 1):
        for r in range(1, output_ws.max_row + 1):
            cell_value = output_ws.cell(row=r, column=c).value
            if cell_value in mappedDict:
                output_ws.cell(row=r, column=c + 4).value = mappedDict[cell_value][0]
                output_ws.cell(row=r, column=c + 5).value = mappedDict[cell_value][1]
            if cell_value == "Total" and (r >= (7 + row_offset) and r <= (7 + row_offset + len(headers))):
                output_ws.cell(row=r, column=c + 4).value = sumOfcurr
                output_ws.cell(row=r, column=c + 5).value = sumofprev

    # Apply thick border to the row immediately after the section
    for col in range(3, 9):  # C to H
        cell = output_ws.cell(row=7 + row_offset + len(headers) , column=col)
        cell.border = thin_border
    cell1=output_ws[f'G{5 + row_offset}']
    cell1.border=None
    input_wb.close()
#-----------------------------------------------End of SubHead--------------------------------------------------
#-----------------------------------------------Nature----------------------------------------------------------
def add_section_with_nature(headers, search_head, note, category,nature, row_offset):
    # Select the appropriate worksheet based on category
    output_ws = bs_ws if category == 'BS' else pl_ws

    # Load the input Excel file
    input_wb = openpyxl.load_workbook(inputPath)
    input_ws = input_wb.active

    # Set the headers
    output_ws[f'C{6 + row_offset}'] = note
    output_ws[f'H{5 + row_offset}'] = "In Rs. hundreds"
    output_ws[f'C{7 + row_offset + len(headers)}'] = "Total"

    output_ws[f'G{6 + row_offset}'] = dateCurr
    output_ws[f'H{6 + row_offset}'] = datePrev

    header_font = Font(size=10, bold=True)
    output_ws[f'G{5 + row_offset}'].font = header_font
    output_ws[f'G{6 + row_offset}'].font = header_font
    output_ws[f'H{6 + row_offset}'].font = header_font
    output_ws[f'C{6 + row_offset}'].font = header_font
    output_ws[f'C{7 + row_offset + len(headers)}'].font = header_font
    output_ws[f'G{7 + row_offset + len(headers)}'].font = header_font
    output_ws[f'H{7 + row_offset + len(headers)}'].font = header_font

    # Apply the border to the header cells
    # for cell in [f'C{6 + row_offset}', f'G{5 + row_offset}', f'G{6 + row_offset}', f'H{6 + row_offset}']:
    #     output_ws[cell].border = thin_border

    # Place each header in a separate row
    for i, header in enumerate(headers):
        output_ws[f'C{7 + row_offset + i}'] = header
        # output_ws[f'C{7 + row_offset + i}'].border = thin_border

    # Specify the header name of the column you want to iterate through
    header_name = "Major Head"
    sub_head_list = []
    mapped_dict = {}

    # Find the index of the column with the specified header name
    column_index = None
    for col in range(1, input_ws.max_column + 1):
        if input_ws.cell(row=1, column=col).value == header_name:
            column_index = col
            break

    if column_index is not None:
        # Iterate through each row in the column
        for row in range(2, input_ws.max_row + 1):  # Start from 2 to skip the header row
            cell_value = input_ws.cell(row=row, column=column_index).value
            if cell_value == search_head:
                sub_head_list.append(input_ws.cell(row=row, column=column_index + 1).value)

        for row in range(2, input_ws.max_row + 1):
            sub_head_value = input_ws.cell(row=row, column=column_index + 1).value
            curr=input_ws.cell(row=row, column=column_index - 2).value,
            prev=input_ws.cell(row=row, column=column_index - 1).value,
            nat=input_ws.cell(row=row, column=column_index + 2).value #capture nature 
            if sub_head_value in sub_head_list:
                if(sub_head_value in mapped_dict):
                    mapped_dict[sub_head_value][0]+=curr
                    mapped_dict[sub_head_value][1]+=prev
                else :
                    mapped_dict[sub_head_value] = [ curr, prev,nat]

    sum_of_curr = sum(mapped_dict[key][0] for key in mapped_dict if mapped_dict[key][0] is not None)
    sum_of_prev = sum(mapped_dict[key][1] for key in mapped_dict if mapped_dict[key][1] is not None)

    for c in range(1, output_ws.max_column + 1):
        for r in range(1, output_ws.max_row + 1):
            cell_value = output_ws.cell(row=r, column=c).value
            if cell_value in mapped_dict:
                output_ws.cell(row=r, column=c + 4).value = mapped_dict[cell_value][0]
                output_ws.cell(row=r, column=c + 5).value = mapped_dict[cell_value][1]

            if cell_value == "Total" and (r >= (7 + row_offset) and r <= (7 + row_offset + len(headers))):
                output_ws.cell(row=r, column=c + 4).value = sum_of_curr
                output_ws.cell(row=r, column=c + 5).value = sum_of_prev

    # Apply thick border to the row immediately after the section
    for col in range(3, 9):  # C to H
        cell = output_ws.cell(row=7 + row_offset + len(headers), column=col)
        # cell.border = thin_border
    cell1 = output_ws[f'G{5 + row_offset}']
    cell1.border = None
    input_wb.close()
#-----------------------------------------------End of Nature---------------------------------------------------
#-----------------------------------------------Nature Division Function ---------------------------------------
def add_section_with_Nature_div(headers, search_head, note, category, subtosub, row_offset,nature1,nature2):
    # Select the appropriate worksheet based on category
    output_ws = bs_ws if category == 'BS' else pl_ws

    # Load the input Excel file
    input_wb = load_workbook(inputPath)
    input_ws = input_wb.active

    # Set the headers
    output_ws[f'C{6 + row_offset}'] = note
    output_ws[f'C{6 + row_offset}'].font=Font(bold=True)
    output_ws[f'H{5 + row_offset}'] = "In Rs. hundreds"
    output_ws[f'C{7 + row_offset}'] = nature1
    output_ws[f'C{7 + row_offset}'].font = Font(bold=True)
    output_ws[f'C{7 + row_offset + len(headers[0]) + 1}'] = nature2
    output_ws[f'C{7 + row_offset + len(headers[0]) + 1}'].font = Font(bold=True)
    output_ws[f'C{7 + row_offset + len(headers[0]) + len(headers[1]) + 2}'] = "Total"
    output_ws[f'C{7 + row_offset + len(headers[0]) + len(headers[1]) + 2}'].font = Font(bold=True)
    output_ws[f'G{6 + row_offset}'] = dateCurr
    output_ws[f'G{6 + row_offset}'].font = Font(bold=True)
    output_ws[f'H{6 + row_offset}'] = datePrev
    output_ws[f'H{6 + row_offset}'].font = Font(bold=True)
    output_ws[f'G{7 + row_offset + len(headers[0]) + len(headers[1]) + 2}'].font = Font(bold=True)
    output_ws[f'H{7 + row_offset + len(headers[0]) + len(headers[1]) + 2}'].font = Font(bold=True)

    # Specify the header name of the column you want to iterate through
    header_name = "Major Head"
    Seq_subhead_list = []
    Unseq_subhead_list = []
    mapped_dict_seq = {}
    mapped_dict_Unseq = {}

    # Set border style
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Set headers
    for i, header in enumerate(headers[0]):
        output_ws[f'C{7 + row_offset + i + 1}'].value = header
        # output_ws[f'C{7 + row_offset + i + 1}'].border = thin_border

    j = len(headers[0]) + 2
    for i, header in enumerate(headers[1]):
        output_ws[f'C{7 + row_offset + j + i}'].value = header
        # output_ws[f'C{7 + row_offset + j + i}'].border = thin_border

    # Find the index of the column with the specified header name
    column_index = None
    for col in range(1, input_ws.max_column + 1):
        if input_ws.cell(row=1, column=col).value == header_name:
            column_index = col
            break
    if(nature1=="Secured"):
        if column_index is not None:
            for row in range(2, input_ws.max_row + 1):  # Start from 2 to skip the header row
                cell_value = input_ws.cell(row=row, column=column_index).value
                if cell_value == search_head:
                    if input_ws.cell(row=row, column=column_index + 3).value == nature1:
                        Seq_subhead_list.append(input_ws.cell(row=row, column=column_index + 1).value)
                    if input_ws.cell(row=row, column=column_index + 3).value == nature2:
                        Unseq_subhead_list.append(input_ws.cell(row=row, column=column_index + 1).value)

            for row in range(2, input_ws.max_row + 1):
                sub_head_value = input_ws.cell(row=row, column=column_index + 1).value
                nature_value = input_ws.cell(row=row, column=column_index + 3).value
                amount_curr = input_ws.cell(row=row, column=column_index - 2).value or 0
                amount_prev = input_ws.cell(row=row, column=column_index - 1).value or 0

                if sub_head_value in Seq_subhead_list and nature_value == nature1:
                    if sub_head_value in mapped_dict_seq:
                        mapped_dict_seq[sub_head_value][0] += amount_curr
                        mapped_dict_seq[sub_head_value][1] += amount_prev
                    else:
                        mapped_dict_seq[sub_head_value] = [amount_curr, amount_prev, nature_value]

                if sub_head_value in Unseq_subhead_list and nature_value == nature2:
                    if sub_head_value in mapped_dict_Unseq:
                        mapped_dict_Unseq[sub_head_value][0] += amount_curr
                        mapped_dict_Unseq[sub_head_value][1] += amount_prev
                    else:
                        mapped_dict_Unseq[sub_head_value] = [amount_curr, amount_prev, nature_value]

        print("Secured list:", Seq_subhead_list)
        print("Unsecured list:", Unseq_subhead_list)
        print("Mapped Dict Seq:", mapped_dict_seq)
        print("Mapped Dict Unseq:", mapped_dict_Unseq)


        sum_of_curr = sum(mapped_dict_seq[key][0] for key in mapped_dict_seq if mapped_dict_seq[key][0] is not None)+sum(mapped_dict_Unseq[key][0] for key in mapped_dict_Unseq if mapped_dict_Unseq[key][0] is not None)
        sum_of_prev = sum(mapped_dict_seq[key][0] for key in mapped_dict_seq if mapped_dict_seq[key][1] is not None)+sum(mapped_dict_Unseq[key][1] for key in mapped_dict_Unseq if mapped_dict_Unseq[key][1] is not None)
        # print(sum_of_curr)
        # print(sum_of_prev)
        for c in range(1, output_ws.max_column + 1):
            for r in range(1, output_ws.max_row + 1):
                cell_value = output_ws.cell(row=r, column=c).value
                if(cell_value==nature1):
                    for r_secured in range(r + 1, output_ws.max_row + 1):
                        cell_value_secured = output_ws.cell(row=r_secured, column=c).value
                        #print(cell_value_secured)
                        if(cell_value_secured!=nature2):
                            if(cell_value_secured in mapped_dict_seq):
                                output_ws.cell(row=r_secured, column=c+4).value=mapped_dict_seq[cell_value_secured][0]
                                output_ws.cell(row=r_secured, column=c+5).value=mapped_dict_seq[cell_value_secured][1]
                        else:
                            break
                if(cell_value==nature2):
                    for r_unseq in range(r + 1, output_ws.max_row + 1):
                        cell_value_unseq = output_ws.cell(row=r_unseq, column=c).value
                        #print(cell_value_unseq)
                        if(cell_value_unseq!="Total"):
                            if(cell_value_unseq in mapped_dict_Unseq):
                                output_ws.cell(row=r_unseq, column=c+4).value=mapped_dict_Unseq[cell_value_unseq][0]
                                output_ws.cell(row=r_unseq, column=c+5).value=mapped_dict_Unseq[cell_value_unseq][1]
                        else:
                            break
        output_ws[f'G{7 + row_offset + len(headers[0]) + len(headers[1]) + 2}'].value=sum_of_curr
        output_ws[f'H{7 + row_offset + len(headers[0]) + len(headers[1]) + 2}'].value=sum_of_prev
        # Calculate the number of rows added in this section
        rows_added = len(headers) + len(headers[0]) + len(headers[1]) + 3  # Adjust 6 as per total rows added including titles

        input_wb.close()

        return row_offset + rows_added
    if(nature1=="Raw Material Consumed"):

        raw_list=[]
        store_list=[]
        mapped_dict_raw={}
        mapped_dict_store={}
        
        if column_index is not None:
            for row in range(2, input_ws.max_row + 1):  # Start from 2 to skip the header row
                cell_value = input_ws.cell(row=row, column=column_index).value
                if cell_value == search_head:
                    if input_ws.cell(row=row, column=column_index + 1).value == nature1:
                        raw_list.append(input_ws.cell(row=row, column=column_index + 1).value)
                    if input_ws.cell(row=row, column=column_index + 1).value == nature2:
                        store_list.append(input_ws.cell(row=row, column=column_index + 1).value)

            for row in range(2, input_ws.max_row + 1):
                sub_head_value = input_ws.cell(row=row, column=column_index + 1).value
                sub_to_sub = input_ws.cell(row=row, column=column_index + 2).value
                amount_curr = input_ws.cell(row=row, column=column_index - 2).value or 0
                amount_prev = input_ws.cell(row=row, column=column_index - 1).value or 0

                if sub_head_value in raw_list and sub_to_sub:
                    #print(sub_head_value)
                    if sub_head_value in mapped_dict_raw:
                        mapped_dict_raw[sub_head_value][0] += amount_curr
                        mapped_dict_raw[sub_head_value][1] += amount_prev
                    else:
                        mapped_dict_raw[sub_head_value] = [amount_curr, amount_prev, sub_to_sub]

                if sub_head_value in store_list and sub_to_sub:
                    if sub_head_value in mapped_dict_store:
                        mapped_dict_store[sub_head_value][0] += amount_curr
                        mapped_dict_store[sub_head_value][1] += amount_prev
                    else:
                        mapped_dict_store[sub_head_value] = [amount_curr, amount_prev, sub_to_sub]

        print("raw list:", raw_list)
        print("store list:", store_list)
        print("Mapped Dict raw:", mapped_dict_raw)
        print("Mapped Dict store:", mapped_dict_store)


        sum_of_curr = sum(mapped_dict_raw[key][0] for key in mapped_dict_raw if mapped_dict_raw[key][0] is not None)+sum(mapped_dict_store[key][0] for key in mapped_dict_store if mapped_dict_store[key][0] is not None)
        sum_of_prev = sum(mapped_dict_raw[key][1] for key in mapped_dict_raw if mapped_dict_raw[key][1] is not None)+sum(mapped_dict_store[key][1] for key in mapped_dict_store if mapped_dict_store[key][1] is not None)
        # print(sum_of_curr)
        # print(sum_of_prev)
        for c in range(1, output_ws.max_column + 1):
            for r in range(1, output_ws.max_row + 1):
                cell_value = output_ws.cell(row=r, column=c).value
                if(cell_value==nature1):
                    for r_secured in range(r + 1, output_ws.max_row + 1):
                        cell_value_secured = output_ws.cell(row=r_secured, column=c).value
                        #print(cell_value_secured)
                        if(cell_value_secured!=nature2):
                            # if(cell_value_secured in mapped_dict_raw or (cell_value_secured in mapped_dict_raw and cell_value_secured == mapped_dict_raw[cell_value_secured][2])):
                            #     output_ws.cell(row=r_secured, column=c+4).value=mapped_dict_raw[cell_value_secured][0]
                            #     output_ws.cell(row=r_secured, column=c+5).value=mapped_dict_raw[cell_value_secured][1]
                            for k,v in mapped_dict_raw.items():
                                if(cell_value_secured==v[2]):
                                    output_ws.cell(row=r_secured, column=c+4).value=mapped_dict_raw[nature1][0]
                                    output_ws.cell(row=r_secured, column=c+5).value=mapped_dict_raw[nature1][1]
                        else:
                            break
                if(cell_value==nature2):
                    for r_unseq in range(r + 1, output_ws.max_row + 1):
                        cell_value_unseq = output_ws.cell(row=r_unseq, column=c).value
                        #print(cell_value_unseq)
                        if(cell_value_unseq!="Total"):
                            # if( cell_value_unseq in mapped_dict_store and cell_value_unseq == mapped_dict_store[cell_value_unseq][2]):
                            #     output_ws.cell(row=r_unseq, column=c+4).value=mapped_dict_store[cell_value_unseq][0]
                            #     output_ws.cell(row=r_unseq, column=c+5).value=mapped_dict_store[cell_value_unseq][1]
                            for k,v in mapped_dict_store.items():
                                if(cell_value_unseq==v[2]):
                                    output_ws.cell(row=r_unseq, column=c+4).value=mapped_dict_store[nature2][0]
                                    output_ws.cell(row=r_unseq, column=c+5).value=mapped_dict_store[nature2][1]

                        else:
                            break
        output_ws[f'G{7 + row_offset + len(headers[0]) + len(headers[1]) + 2}'].value=sum_of_curr
        output_ws[f'H{7 + row_offset + len(headers[0]) + len(headers[1]) + 2}'].value=sum_of_prev
        # Calculate the number of rows added in this section
        rows_added = len(headers) + len(headers[0]) + len(headers[1]) + 3  # Adjust 6 as per total rows added including titles

        input_wb.close()

        return row_offset + rows_added
#-----------------------------------------------end of Nature-Division Function --------------------------------
#-----------------------------------------------Trade Template Function Only------------------------------------
def add_template_with_intervals(ws, headers, search_head, note, category, intervals,date, row_offset):
    # Load the input Excel file
    input_wb = openpyxl.load_workbook(inputPath)
    input_ws = input_wb.active

    # Set the note heading
    ws[f'C{6 + row_offset}'] = note
    ws[f'C{6 + row_offset}'].font = Font(bold=True)
    # Set the column headers for intervals
    col_headings = [date] + intervals if isinstance(dateCurr, str) else [dateCurr] + intervals
    for i, heading in enumerate(col_headings):
       cellval= ws.cell(row=7 + row_offset, column=3 + i)
       cellval.value=heading
       cellval.font=Font(bold=True)
       cellval.alignment= Alignment(horizontal='center')


    # Set the border for column headers
    # for col in range(3, 3 + len(col_headings)):  # Columns C to G
    #     ws.cell(row=7 + row_offset, column=col).border = thin_border

    # Set the subhead list
    subhead_row = 8 + row_offset
    for i, header in enumerate(headers):
        ws.cell(row=subhead_row + i, column=3).value = header
        # ws.cell(row=subhead_row + i, column=3).border = thin_border
    column_index = None
    header_name='Major Head'
    subHeadList=[]
    mappedDict=[]
    for col in range(1, input_ws.max_column + 1):
        if input_ws.cell(row=1, column=col).value == header_name:
            column_index = col
            break
    if(column_index is not None):
        for row in range(2, input_ws.max_row + 1):  # Start from 2 to skip the header row
            cell_value = input_ws.cell(row=row, column=column_index).value
            if cell_value == search_head:
                subHeadList.append(input_ws.cell(row=row, column=column_index + 1).value)
        for row in range(2, input_ws.max_row + 1):
            sub_head_value = input_ws.cell(row=row, column=column_index + 1).value
            currYearval=input_ws.cell(row=row, column=column_index - 2).value or 0 
            prevYearval=input_ws.cell(row=row, column=column_index - 1).value or 0
            if sub_head_value in subHeadList:
                if(sub_head_value in mappedDict):
                    mappedDict[sub_head_value][0]+=currYearval
                    mappedDict[sub_head_value][1]+=prevYearval
                else :
                    mappedDict[sub_head_value] = [currYearval,prevYearval]

    sumOfcurr = sum(mappedDict[key][0] for key in mappedDict if mappedDict[key][0] is not None)
    sumofprev = sum(mappedDict[key][1] for key in mappedDict if mappedDict[key][1] is not None)
        
    # Set the "Total" row
    total_row = subhead_row + len(headers)
    ws.cell(row=total_row, column=3).value = "Total"
    ws.cell(row=total_row, column=7).value = sumOfcurr
    ws.cell(row=total_row, column=8).value = sumofprev
    ws.cell(row=total_row, column=3).border = thin_border
    ws.cell(row=total_row, column=3).font=Font(bold=True)
    ws.cell(row=total_row, column=4).font=Font(bold=True)
    ws.cell(row=total_row, column=5).font=Font(bold=True)
    ws.cell(row=total_row, column=6).font=Font(bold=True)
    ws.cell(row=total_row, column=7).font=Font(bold=True)
    ws.cell(row=total_row, column=8).font=Font(bold=True)
    input_wb.close()
#-----------------------------------------------end of add_template_with_intervals------------------------------
def add_section_with_Nature_div_2(headers, search_head, note, category, subtosub, row_offset, nature1, nature2, nature3):
    # Select the appropriate worksheet based on category
    output_ws = bs_ws if category == 'BS' else pl_ws

    # Load the input Excel file
    input_wb = load_workbook(inputPath)
    input_ws = input_wb.active

    # Set the headers
    output_ws[f'C{6 + row_offset}'] = note
    output_ws[f'C{6 + row_offset}'].font = Font(bold=True)
    output_ws[f'H{5 + row_offset}'] = "In Rs. hundreds"
    output_ws[f'C{7 + row_offset}'] = nature1
    output_ws[f'C{7 + row_offset}'].font = Font(bold=True)
    output_ws[f'C{7 + row_offset + len(headers[0]) + 1}'] = nature2
    output_ws[f'C{7 + row_offset + len(headers[0]) + 1}'].font = Font(bold=True)
    output_ws[f'C{7 + row_offset + len(headers[0]) + len(headers[1]) + len(headers[2]) + 3}']="Total"
    output_ws[f'C{7 + row_offset + len(headers[0]) + len(headers[1]) + len(headers[2]) + 3}'].font=Font(bold=True)
    output_ws[f'G{7 + row_offset + len(headers[0]) + len(headers[1]) + len(headers[2]) + 3}'].font=Font(bold=True)
    output_ws[f'H{7 + row_offset + len(headers[0]) + len(headers[1]) + len(headers[2]) + 3}'].font=Font(bold=True)
    output_ws[f'G{6 + row_offset}'] = dateCurr
    output_ws[f'G{6 + row_offset}'].font = Font(bold=True)
    output_ws[f'H{6 + row_offset}'] = datePrev
    output_ws[f'H{6 + row_offset}'].font = Font(bold=True)
    output_ws[f'C{7 + row_offset + len(headers[0]) + len(headers[1]) + 2}'] = nature3
    output_ws[f'C{7 + row_offset + len(headers[0]) + len(headers[1]) + 2}'].font = Font(bold=True)

    # Specify the header name of the column you want to iterate through
    header_name = "Major Head"
    finish_list = []
    progress_list = []
    stock_list = []
    mapped_dict_finish = {}
    mapped_dict_progress = {}
    mapped_dict_stock = {}

    # Set border style

    # Set headers
    for i, header in enumerate(headers[0]):
        output_ws[f'C{7 + row_offset + i + 1}'].value = header
        # output_ws[f'C{7 + row_offset + i + 1}'].border = thin_border

    j = len(headers[0]) + 2
    for i, header in enumerate(headers[1]):
        output_ws[f'C{7 + row_offset + j + i}'].value = header
        # output_ws[f'C{7 + row_offset + j + i}'].border = thin_border

    k = len(headers[0]) + len(headers[1]) + 3
    for i, header in enumerate(headers[2]):
        output_ws[f'C{7 + row_offset + k + i}'].value = header
        # output_ws[f'C{7 + row_offset + k + i}'].border = thin_border

    # Find the index of the column with the specified header name
    column_index = None
    for col in range(1, input_ws.max_column + 1):
        if input_ws.cell(row=1, column=col).value == header_name:
            column_index = col
            break
    
    if nature1:
        if column_index is not None:
            for row in range(2, input_ws.max_row + 1):  # Start from 2 to skip the header row
                cell_value = input_ws.cell(row=row, column=column_index).value
                if cell_value == search_head:
                    if input_ws.cell(row=row, column=column_index + 3).value == nature1:
                        finish_list.append(input_ws.cell(row=row, column=column_index + 1).value)
                    if input_ws.cell(row=row, column=column_index + 3).value == nature2:
                        progress_list.append(input_ws.cell(row=row, column=column_index + 1).value)
                    if input_ws.cell(row=row, column=column_index + 3).value == nature3:
                        stock_list.append(input_ws.cell(row=row, column=column_index + 1).value)

            for row in range(2, input_ws.max_row + 1):
                sub_head_value = input_ws.cell(row=row, column=column_index + 1).value
                sub_to_sub = input_ws.cell(row=row, column=column_index + 2).value
                amount_curr = input_ws.cell(row=row, column=column_index - 2).value or 0
                amount_prev = input_ws.cell(row=row, column=column_index - 1).value or 0

                if sub_head_value in finish_list and sub_to_sub == nature1:
                    if sub_head_value in mapped_dict_finish:
                        mapped_dict_finish[sub_head_value][0] += amount_curr
                        mapped_dict_finish[sub_head_value][1] += amount_prev
                    else:
                        mapped_dict_finish[sub_head_value] = [amount_curr, amount_prev, sub_to_sub]

                if sub_head_value in progress_list and sub_to_sub == nature2:
                    if sub_head_value in mapped_dict_progress:
                        mapped_dict_progress[sub_head_value][0] += amount_curr
                        mapped_dict_progress[sub_head_value][1] += amount_prev
                    else:
                        mapped_dict_progress[sub_head_value] = [amount_curr, amount_prev, sub_to_sub]
                
                if sub_head_value in stock_list and sub_to_sub == nature3:
                    if sub_head_value in mapped_dict_stock:
                        mapped_dict_stock[sub_head_value][0] += amount_curr
                        mapped_dict_stock[sub_head_value][1] += amount_prev
                    else:
                        mapped_dict_stock[sub_head_value] = [amount_curr, amount_prev, sub_to_sub]

        print("finish list:", finish_list)
        print("progress list:", progress_list)
        print("stock list:", stock_list)
        print("Mapped Dict finish:", mapped_dict_finish)
        print("Mapped Dict progress:", mapped_dict_progress)
        print("Mapped Dict stock:", mapped_dict_stock)

        sum_of_curr = sum(mapped_dict_finish[key][0] for key in mapped_dict_finish if mapped_dict_finish[key][0] is not None) + \
                      sum(mapped_dict_progress[key][0] for key in mapped_dict_progress if mapped_dict_progress[key][0] is not None) + \
                      sum(mapped_dict_stock[key][0] for key in mapped_dict_stock if mapped_dict_stock[key][0] is not None)
        
        sum_of_prev = sum(mapped_dict_finish[key][1] for key in mapped_dict_finish if mapped_dict_finish[key][1] is not None) + \
                      sum(mapped_dict_progress[key][1] for key in mapped_dict_progress if mapped_dict_progress[key][1] is not None) + \
                      sum(mapped_dict_stock[key][1] for key in mapped_dict_stock if mapped_dict_stock[key][1] is not None)
        
        # Update output worksheet with the calculated sums and mapped values
        for c in range(1, output_ws.max_column + 1):
            for r in range(row_offset, row_offset + len(headers[0]) + len(headers[1]) + 2):
                cell_value = output_ws.cell(row=r, column=c).value
                if cell_value == nature1:
                    for r_secured in range(r + 1, row_offset + len(headers[0]) + len(headers[1]) + 2):
                        cell_value_secured = output_ws.cell(row=r_secured, column=c).value
                        if cell_value_secured != nature2:
                            if cell_value_secured in mapped_dict_finish:
                                for k, v in mapped_dict_finish.items():
                                    if cell_value_secured == v[2]:
                                        output_ws.cell(row=r_secured, column=c + 4).value = mapped_dict_finish[cell_value_secured][0]
                                        output_ws.cell(row=r_secured, column=c + 5).value = mapped_dict_finish[cell_value_secured][1]
                        else:
                            break
                if cell_value == nature2:
                    for r_unseq in range(r + 1, row_offset + len(headers[0]) + len(headers[1]) + 2):
                        cell_value_unseq = output_ws.cell(row=r_unseq, column=c).value
                        if cell_value_unseq != nature3:
                            if cell_value_unseq in mapped_dict_progress:
                                for k, v in mapped_dict_finish.items():
                                    if cell_value_unseq == v[2]:
                                        output_ws.cell(row=r_unseq, column=c + 4).value = mapped_dict_progress[cell_value_unseq][0]
                                        output_ws.cell(row=r_unseq, column=c + 5).value = mapped_dict_progress[cell_value_unseq][1]
                        else:
                            break
                if cell_value == nature3:
                    for r_stock in range(r + 1, row_offset + len(headers[0]) + len(headers[1]) + len(headers[2]) + 2):
                        cell_value_stock = output_ws.cell(row=r_stock, column=c).value
                        if cell_value_stock != "Total":
                            if cell_value_stock in mapped_dict_stock:
                                for k, v in mapped_dict_stock.items():
                                    if cell_value_stock == v[2]:
                                        output_ws.cell(row=r_stock, column=c + 4).value = mapped_dict_stock[cell_value_stock][0]
                                        output_ws.cell(row=r_stock, column=c + 5).value = mapped_dict_stock[cell_value_stock][1]
                        else:
                            break
        output_ws[f'G{7 + row_offset + len(headers[0]) + len(headers[1]) + len(headers[2]) + 3}'].value = sum_of_curr
        output_ws[f'H{7 + row_offset + len(headers[0]) + len(headers[1]) + len(headers[2]) + 3}'].value = sum_of_prev
        
        # Calculate the number of rows added in this section
        rows_added = len(headers[0]) + len(headers[1]) + len(headers[2]) + 8  # Adjust to include all rows

        input_wb.close()

        return row_offset + rows_added


#List of ALL the major Heads and it Sub Heads 
otherCurrentLiabilities = [
    "Current Maturities Of Finance Lease Obligations",
    "Interest Accrued But Not Due On Borrowings",
    "Interest Accrued And Due On Borrowings",
    "Income Received In Advance",
    "Unpaid Dividends",
    "Application Money Received For Allotment Of Securities And Due For Refund And Interest Accrued Thereon",
    "Unpaid Matured Deposits And Interest Accrued Thereon",
    "Unpaid Matured Debentures And Interest Accrued Thereon",
    "Other Payables (Specify Nature)"
]
DeferredTaxLiability = ["Deferred Tax Liability"]
longtermliabilities = ["Trade Payables", "Others"]
cashandcashequivalent = [
    "Cheques, Drafts On Hand",
    "Cash On Hand",
    "Cash At Bank",
    "Other C & CE (Specify Nature)"
]
CurrentInvestments = [
    "Investments In Equity Instruments",
    "Investments In Preference Shares",
    "Investments In Government Or Trust Securities",
    "Investments In Debentures Or Bonds",
    "Investments In Mutual Funds",
    "Investments In Partnership Firms",
    "Other Investments (Specify Nature)"
]
LongtermProvisions = [
    "Provision For Employee Benefits",
    "Others Long Term Provision (Specify Nature)"
]
shortTermProvisions = [
    "Provision for Employees Benefits",
    "Provision for Income Tax",
    "Other Short Term Provision (Specify Nature)"
]
nonCurrentInvestments = [
    "Investment Property",
    "Investments In Equity Instruments",
    "Investments In Preference Shares",
    "Investments In Government Or Trust Securities",
    "Investments In Debentures Or Bonds",
    "Investments In Mutual Funds",
    "Investments In Partnership Firms",
    "Other Non-Current Investments (Specify Nature)"
]
deferredtaxassets = ["Deferred Tax Asset"]
inventories = [
    "Raw Materials",
    "Work-In-Progress",
    "Finished Goods",
    "Stock-In-Trade",
    "Stores And Spares",
    "Loose Tools",
    "Other Inventory (Specify Nature)"
]
othercurrentAssets=[
    "Other Current Assets 1",
    "Other Current Assets 2",
    "Other Current Assets 3",
    "Other Current Assets 4",
    "Other Current Assets 5"
]
revenueFromOperations=[
    "Sale Of Product",
    "Sale of Services",
    "Grants or Donations Received",
    "Other Operating Revenue",
    "Less: Excise Duty"
]
otherIncome=[
    "Interest Income",
    "Dividend Income",
    "Net Gain/Loss On Sale Of Investments",
    "Other Non - Operating Income"
]
purchaseofstockintrade=["Purchases"]
EmployeeRem=[
    "Salary & Wages",
    "Employees State Insurance",
    "Employers Contribution To Providend Fund",
    "Director Remuneration"
]
financecost=[
    "Interest Paid During The Year",
    "Other Borrowing Cost",
    "Applicable Net Gain/Loss On Foreign Currency Transactions And Translation"
]
exceptionItems=["Exceptional Item 1","Exceptional Item 2"]
extraordinaryItems=["Extraordinary Item 1","Extraordinary Item 2"]
longTermLossesAndAdvances = [
    "Secured",
    "Unsecured",
    "Doubtful",
    "Loans And Advances To Related Parties",
    "Other Long Term Loans And Advances (Specify Nature)"
]
othernoncurrentassets=[
    "Long-term Trade Receivables (including trade receivables on deferred credit terms)",
    "Secured, considered good",
    "Unsecured, considered good",
    "Doubtful",
    "Security Deposits",
    "Other Assets"
]
shortTermLoansansAdvances=[
    "Loans And Advances To Related Parties",
    "Secured",
    "Unsecured",
    "Doubtful",
    "Other Short Term Loans and Advances (Specify Nature)"
]
otherExpenses=[
    "1)Payment To Auditors",
    "Statutory Audit",
    "Taxation Matters",
    "2)Other Expenses",
    "Power and Fuel",
    "Labour Contract Charges",
    "Factory Rent",
    "Travelling Expenses",
    "Misc. Expenses",
    "Repair To Machinery",
    "Rates And Taxes",
    "Office Expenses",
    "Operating Expenses",
    "Other Repair And Maintenance",
    "Insurance",
    "Repairs To Building",
    "Printing Expenses",
    "Debt Written Off",
    "Interest On Income Tax"
]
longTermBorrowings= [[
    "Bonds and Debentures",
    "Term Loans",
    "Term Loans from Banks",
    "Term Loans from Other Parties",
    "Deposits"],
    [
    "Bonds and Debentures",
    "Term Loans from Banks",
    "Term Loans from Other Parties",
    "Deposits",
    "Deferred Payment Liabilities",
    "Loans and Advances From Related Parties",
    "Long Term Maturities Of Finance Lease Obligations",
    "Other Loans And Advances (Specify Nature)"]
]
tradePayables=[
    "MSME",
    "Others",
    "Disputed Dues - MSME",
    "Disputed Dues - Others"
]
time_intervals_1 = [
    "less than 1 Year",
    "1 year - 2 years",
    "2 years - 3 years",
    "more than 3 years",
    "Total"
]
tradeRecieveables=[
    "Undisputed Trade receivables - considered good",
    "Undisputed Trade Receivables - considered doubtful",
    "Disputed Trade Receivables considered good",
    "Disputed Trade Receivables considered doubtful"
]
time_intervals_2 = [
    "less than 6 months",
    "6 months - 1 year",
    "1 year - 2 years",
    "2 years - 3 years",
    "more than 3 years",
    "Total"
]
shortTermBorrowings=[
    ["Loans Repayable on Demand from Banks","Loans Repayable on Demand from Other Parties","Deposits"],
    [
    "Unsecured Borrowings",
    "Loans Repayable on Demand from Banks","Loans Repayable on Demand from Other Parties",
    "Deposits",
    "Loans And Advances From Related Parties",
    "Current Maturities of Long Term Debt",
    "Other Short Term Loans And Advances (Specify Nature)"
    ]
]
costofmaterialsconsumed=[
    ["Opening Stock",
    "Purchases",
    "Closing Stock",
    "Raw Material Consumed"],
     [
    "Opening Stock",
    "Purchases",
    "Closing Stock",
    "Stores & Consumables"
]
]
IncreaseDecreaseinStock=[
    [ "Opening Stock", "Closing Stock"],
    [ "Opening Stock", "Closing Stock"],
    [ "Opening Stock", "Closing Stock"]
]
sections = [
    {"header": longTermBorrowings, "search_head": "Long-Term Borrowings", "note": "Note : Long-Term Borrowings", "category": "BS", "nature":None, "subtosub": "present", "interval": None},
    {"header": DeferredTaxLiability, "search_head": "Deferred Tax Liability", "note": "Note : Deferred Tax Liability", "category": "BS", "nature": None, "subtosub": None, "interval": None},
    {"header": longtermliabilities, "search_head": "Other Long-term Liabilities", "note": "Note : Other Long-term Liabilities", "category": "BS", "nature": None, "subtosub": None, "interval": None},
    {"header": LongtermProvisions, "search_head": "Long-Term Provisions", "note": "Note : Long-Term Provisions", "category": "BS", "nature": None, "subtosub": None, "interval": None},
    {"header": shortTermBorrowings, "search_head": "Short Term Borrowing", "note": "Note : Short Term Borrowing", "category": "BS", "nature": None, "subtosub":"present", "interval": None},
    {"header": tradePayables, "search_head": "Trade Payables", "note": "Note : Trade Payables", "category": "BS", "nature": None, "subtosub": None, "interval": time_intervals_1},
    {"header": otherCurrentLiabilities, "search_head": "Other Current Liabilities", "note": "Note : Other Current Liabilities", "category": "BS", "nature": None, "subtosub": None, "interval": None},
    {"header": shortTermProvisions, "search_head": "Short Term Provisions", "note": "Note : Short Term Provisions", "category": "BS", "nature": None, "subtosub": None, "interval": None},
    {"header": nonCurrentInvestments, "search_head": "Non Current Investments", "note": "Note : Non Current Investments", "category": "BS", "nature": None, "subtosub": None, "interval": None},
    {"header": deferredtaxassets, "search_head": "Deferred Tax Asset", "note": "Note : Deferred Tax Asset", "category": "BS", "nature": None, "subtosub": None, "interval": None},
    {"header": longTermLossesAndAdvances, "search_head": "Long-Term Loans and  Advances", "note": "Note : Long-Term Loans and  Advances", "category": "BS", "nature": "present", "subtosub": None, "interval": None},
    {"header": othernoncurrentassets, "search_head": "Other Non-Current Assets", "note": "Note : Other Non-Current Assets", "category": "BS", "nature": "present", "subtosub": None, "interval": None},
    {"header": CurrentInvestments, "search_head": "Current Investments", "note": "Note : Current Investments", "category": "BS", "nature": None, "subtosub": None, "interval": None},
    {"header": inventories, "search_head": "Inventories", "note": "Note : Inventories", "category": "BS", "nature": None, "subtosub": None, "interval": None},
    {"header": tradeRecieveables, "search_head": "Trade Receivables", "note": "Note : Trade Receivables", "category": "BS", "nature": None, "subtosub": None, "interval": time_intervals_2},
    {"header": shortTermLoansansAdvances, "search_head": "Short Term Loans and Advances", "note": "Note : Short Term Loans and Advances", "category": "BS", "nature": "present", "subtosub": None, "interval": None},
    {"header": cashandcashequivalent, "search_head": "Cash and Cash Equivalents", "note": "Note : Cash and Cash Equivalents", "category": "BS", "nature": None, "subtosub": None, "interval": None},
    {"header": othercurrentAssets, "search_head": "Other Current Assets", "note": "Note : Other Current Assets", "category": "BS", "nature": None, "subtosub": None, "interval": None},
    {"header": revenueFromOperations, "search_head": "Revenue From Operations", "note": "Note : Revenue From Operations", "category": "PL", "nature": None, "subtosub": None, "interval": None},
    {"header": otherIncome, "search_head": "Other Income", "note": "Note : Other Income", "category": "PL", "nature": None, "subtosub": None, "interval": None},
    {"header": costofmaterialsconsumed, "search_head": "Cost of Materials Consumed", "note": " Note : Cost of Materials Consumed", "category": "PL", "nature":None, "subtosub": "present", "interval": None},
    {"header": purchaseofstockintrade, "search_head": "Purchase Of Stock in Trade", "note": "Note : Purchase Of Stock in Trade", "category": "PL", "nature": None, "subtosub": None, "interval": None},
    {"header": IncreaseDecreaseinStock, "search_head": "(Increase)/Decrease in Stocks", "note": "Note : (Increase)/Decrease in Stocks", "category": "PL", "nature":None, "subtosub": "p", "interval": None},
    {"header": EmployeeRem, "search_head": "Employees Remuneration & Benefits", "note": "Note : Employees Remuneration & Benefits", "category": "PL", "nature": None, "subtosub": None, "interval": None},
    {"header": financecost, "search_head": "Finance Cost", "note": "Note : Finance Cost", "category": "PL", "nature": None, "subtosub": None, "interval": None},
    {"header": otherExpenses, "search_head": "Other Expenses", "note": "Note : Other Expenses", "category": "PL", "nature": None, "subtosub": None, "interval": None},
    {"header": exceptionItems, "search_head": "Exceptional Items", "note": "Note : Exceptional Items", "category": "PL", "nature": None, "subtosub": None, "interval": None},
    {"header": extraordinaryItems, "search_head": "Extraordinary Items", "note": "Note : Extraordinary Items", "category": "PL", "nature": None, "subtosub": None, "interval": None}
]


#Calling Every Function As In an Order as Per the Bs and Pl 
row_offset_bs = 0  # Row offset for BS sheet
row_offset_pl = 0  # Row offset for PL sheet
for section in sections:
    if section["category"] == "BS":
        if(section["nature"] is None and section["interval"]is None and section["subtosub"]is None):
            add_section(section["header"], section["search_head"], section["note"], section["category"], row_offset_bs)
            row_offset_bs += len(section["header"]) + 4  # Adjust row offset for next BS section
        if(section["nature"]is not None and section["interval"]is None and section["subtosub"]is None):
            add_section_with_nature(section["header"], section["search_head"], section["note"], section["category"],section["nature"], row_offset_bs)
            row_offset_bs += len(section["header"]) + 4
        if(section["interval"]and section["subtosub"]is None):  
            date=dateCurr
            add_template_with_intervals(bs_ws,section["header"], section["search_head"], section["note"], section["category"],section["interval"],date, row_offset_bs)
            row_offset_bs += len(section["header"]) + 4
            date=datePrev
            add_template_with_intervals(bs_ws,section["header"], section["search_head"], section["note"], section["category"],section["interval"],date, row_offset_bs)
            row_offset_bs += len(section["header"]) + 4
        if(section["subtosub"]):
            row_updated=add_section_with_Nature_div(section["header"],section["search_head"],section["note"],section["category"],section["subtosub"],row_offset_bs,"Secured","Unsecured")
            row_offset_bs=row_updated
    if section["category"] == "PL":
        if(section["subtosub"]):
            if(section["subtosub"]=="present"):
                row_updated=add_section_with_Nature_div(section["header"],section["search_head"],section["note"],section["category"],section["subtosub"],row_offset_pl,"Raw Material Consumed","Stores & Consumables")
                row_offset_pl=row_updated
            if(section["subtosub"]=="p"):
                row_updated=add_section_with_Nature_div_2(section["header"],section["search_head"],section["note"],section["category"],section["subtosub"],row_offset_pl,"Finished Goods","Work-in-Progress","Stock in Trade")
                row_offset_pl=row_updated
        else :
            add_section(section["header"], section["search_head"], section["note"], section["category"], row_offset_pl)
            row_offset_pl += len(section["header"]) + 4  # Adjust row offset for next BS section

# Function to apply border to the area encompassing first and last non-null values for both worksheets
def apply_border_to_first_last_non_null(ws):
    first_non_null_row = None
    first_non_null_col = None
    last_non_null_row = None
    last_non_null_col = None

    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value is not None:
                if first_non_null_row is None:
                    first_non_null_row = row
                    first_non_null_col = col
                last_non_null_row = row
                last_non_null_col = col

# Define border style
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
thick_border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
no_border = Border(left=Side(style=None), right=Side(style=None), top=Side(style=None), bottom=Side(style=None))

bs_ws.sheet_view.showGridLines = False

pl_ws.sheet_view.showGridLines = False

#total border bottom 
border_style_bot_total = Border(bottom=Side(style='thick'))
start_column = 'C'
end_column = 'H'


#C col left border bs
thick_left_border = Border(left=Side(style='thick'))
start_row = 1
end_row = 190
column = 'C'
for row in range(start_row, end_row + 1):
    cell = bs_ws[f'{column}{row}']
    cell.border = thick_left_border
column = 'D'
for row in range(start_row, end_row + 1):
    cell = bs_ws[f'{column}{row}']
    cell.border = thick_left_border
column = 'D'
for row in range(start_row, end_row + 1):
    cell = pl_ws[f'{column}{row}']
    cell.border = thick_left_border

# h col right border  bs
thick_right_border = Border(right=Side(style='thick'))
start_row = 1
end_row = 190
column = 'H'
for row in range(start_row, end_row + 1):
    cell = bs_ws[f'{column}{row}']
    cell.border = thick_right_border
start_row = 1
end_row = 190
#right border to left f bs
column = 'F'
for row in range(start_row, end_row + 1):
    cell = bs_ws[f'{column}{row}']
    cell.border = thick_right_border 
#right border to left f pl
start_row = 1
end_row = 103
column = 'F'
for row in range(start_row, end_row + 1):
    cell = pl_ws[f'{column}{row}']
    cell.border = thick_right_border

#top border bs
row_number = 1
start_column = 'C'
end_column = 'H'
border_style = Border(
    top=Side(style='thick'),
    left=Side(style='thick'),
    right=Side(style='thick')
)
for col in range(ord(start_column), ord(end_column) + 1):
    cell = bs_ws[f'{chr(col)}{row_number}']
    cell.border = border_style
#top border pl
for col in range(ord(start_column), ord(end_column) + 1):
    cell = pl_ws[f'{chr(col)}{row_number}']
    cell.border = border_style

#c col border pl
start_row = 1
end_row = 103
column='C'
for row in range(start_row, end_row + 1):
    cell = pl_ws[f'{column}{row}']
    cell.border = thick_left_border
# h col border pl 
start_row = 1
end_row = 103
column='H'
for row in range(start_row, end_row + 1):
    cell = pl_ws[f'{column}{row}']
    cell.border = thick_right_border
border_style_bot = Border(
    bottom=Side(style='thick'),
    left=Side(style='thick'),
    right=Side(style='thick')
)

#bottom border Bs
start_column = 'C'
end_column = 'H'
for col in range(ord(start_column), ord(end_column) + 1):
    cell = bs_ws[f'{chr(col)}{190}']
    cell.border = border_style_bot
#bottom border pl
for col in range(ord(start_column), ord(end_column) + 1):
    cell = pl_ws[f'{chr(col)}{103}']
    cell.border = border_style_bot


# Save the output workbook
output_wb.save(outputPath)  