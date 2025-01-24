import openpyxl
import pathlib
import time 


def main():
    ip = input("Enter the path of the Trial Balance with '\\': ")
    
    try:
        # Load the Excel file
        workbook = openpyxl.load_workbook(ip)
        sheet = workbook.active

        # Initialize lists to store values
        current_year_values_dr = []
        current_year_values_cr = []
        prev_year_values_dr = []
        prev_year_values_cr = []

        # Iterate through rows and columns
        for row in range(2, sheet.max_row + 1):  # Iterate over rows starting from row 2
            dr_cr_value = sheet.cell(row=row, column=2).value  # Dr/Cr column
            current_year = sheet.cell(row=row, column=3).value
            prev_year = sheet.cell(row=row, column=4).value

            # Process Current year column
            if current_year is not None and isinstance(current_year, (int, float)):
                if dr_cr_value == 'Cr.':
                    current_year_values_cr.append(current_year)
                elif dr_cr_value == 'Dr.':
                    current_year_values_dr.append(current_year)

            # Process Previous year column
            if prev_year is not None and isinstance(prev_year, (int, float)):
                if dr_cr_value == 'Cr.':
                    prev_year_values_cr.append(prev_year)
                elif dr_cr_value == 'Dr.':
                    prev_year_values_dr.append(prev_year)

        print("Sum of Credits Current Year = ", sum(current_year_values_cr))
        print("Sum of Debits Current Year = ", sum(current_year_values_dr))
        if sum(current_year_values_cr) == sum(current_year_values_dr):
            print("Status Matched")
        else:
            print("Status - Not Matched")

        print("Sum of Credits Previous Year = ", sum(prev_year_values_cr))
        print("Sum of Debits Previous Year = ", sum(prev_year_values_dr))
        if sum(prev_year_values_cr) == sum(prev_year_values_dr):
            print("Status Matched")
        else:
            print("Status - Not Matched")

    except Exception as e:
        print("An error occurred: ", e)
    xyz=input("Press any key to exit")
if __name__ == "__main__":
    main()

